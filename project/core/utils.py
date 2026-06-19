
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime

def generate_loyalty_transfer_excel(transfers, transfer_type='single'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "نقاط الولاء"
    
    headers = ['رقم العضوية', 'اسم العميل', 'النقاط المحولة', 'تاريخ التحويل']
    header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row, transfer in enumerate(transfers, 2):
        ws.cell(row=row, column=1, value=transfer.customer.customer_id if transfer.customer else '-')
        ws.cell(row=row, column=2, value=transfer.customer.full_name if transfer.customer else '-')
        ws.cell(row=row, column=3, value=transfer.points)
        ws.cell(row=row, column=5, value=transfer.transferred_at.strftime('%Y-%m-%d %H:%M') if transfer.transferred_at else datetime.now().strftime('%Y-%m-%d %H:%M'))
        
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
    
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    return wb


def create_excel_response(wb, filename):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response